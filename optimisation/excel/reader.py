# ============================================================
# LECTURE DU FICHIER DE CONFIGURATION ATHLETE
# optimisation/excel/reader.py
# ============================================================

import sys
import openpyxl
from pathlib import Path
sys.dont_write_bytecode = True

ROOT_DIR = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT_DIR))

from config import EXCEL_CONFIG_PATH, MAGASIN_FALLBACK

# ------------------------------------------------------------
# MAGASINS VALIDES
# ------------------------------------------------------------
MAGASINS_VALIDES = {
    "leclerc" : "leclerc",
    "auchan"  : "auchan",
}


# ------------------------------------------------------------
# HELPERS
# ------------------------------------------------------------
def _get_cell(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == "" or val.lower() in ["aucun", "aucune"]:
            return None
    return val


def _get_float(ws, row, col, defaut=0.0):
    val = _get_cell(ws, row, col)
    try:
        return float(val)
    except (TypeError, ValueError):
        return defaut


def _get_int(ws, row, col, defaut=0):
    val = _get_cell(ws, row, col)
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return defaut


def _get_str(ws, row, col, defaut=""):
    val = _get_cell(ws, row, col)
    return str(val) if val is not None else defaut


def _is_oui(ws, row, col) -> bool:
    val = _get_str(ws, row, col, "Non")
    return val.strip().lower() in ["oui", "yes", "1", "true"]


# ------------------------------------------------------------
# NORMALISATION MAGASIN
# ------------------------------------------------------------
def _normaliser_magasin(valeur: str) -> str:
    """
    Normalise la valeur Excel vers le code interne.
    Ex: "Leclerc" → "leclerc"
        "AUCHAN"  → "auchan"
        ""        → MAGASIN_FALLBACK
    """
    if not valeur:
        print(f"[WARN] Magasin non renseigné → fallback : {MAGASIN_FALLBACK}")
        return MAGASIN_FALLBACK

    valeur_clean = valeur.strip().lower()

    if valeur_clean in MAGASINS_VALIDES:
        return MAGASINS_VALIDES[valeur_clean]

    print(f"[WARN] Magasin inconnu '{valeur}' → fallback : {MAGASIN_FALLBACK}")
    return MAGASIN_FALLBACK


# ------------------------------------------------------------
# LECTURE FEUILLE PROFIL
# ------------------------------------------------------------
def _read_profil(ws) -> dict:
    return {
        # Colonne gauche
        "nom"                    : _get_str  (ws,  4, 2),
        "age"                    : _get_int  (ws,  5, 2),
        "sexe"                   : _get_str  (ws,  6, 2, "M"),
        "taille_cm"              : _get_float(ws,  7, 2),
        "poids_actuel_kg"        : _get_float(ws,  8, 2),
        "poids_cible_kg"         : _get_float(ws,  9, 2),

        # Colonne droite
        "sport"                  : _get_str  (ws,  4, 5),
        "niveau"                 : _get_str  (ws,  5, 5, "Elite"),
        "objectif"               : _get_str  (ws,  6, 5, "Recomposition"),
        "contexte_medical"       : _get_str  (ws,  7, 5),
        "date_debut"             : _get_str  (ws,  8, 5),

        # Parametres sportifs
        "frequence_entrainement" : _get_int  (ws, 11, 2),
        "duree_seance_min"       : _get_int  (ws, 12, 2),
        "annees_pratique"        : _get_int  (ws, 13, 2),
        "competitions_prevues"   : _get_str  (ws, 14, 2, "Non"),

        # Contraintes alimentaires
        "allergies"              : _get_str  (ws, 16, 2),
        "intolerances"           : _get_str  (ws, 17, 2),
        "aliments_exclus_texte"  : _get_str  (ws, 18, 2),
        "regime_particulier"     : _get_str  (ws, 19, 2),
        "contraintes_digestives" : _get_str  (ws, 20, 2),
    }


# ------------------------------------------------------------
# LECTURE FEUILLE PLANNING
# ------------------------------------------------------------
def _read_planning(ws) -> list:
    planning = []

    for row in range(4, 11):
        jour = _get_str(ws, row, 1)
        if jour:
            planning.append({
                "jour"      : jour,
                "seance"    : _get_str(ws, row, 2, "Repos"),
                "intensite" : _get_str(ws, row, 3, "faible"),
                "duree_min" : _get_int(ws, row, 4, 0),
                "notes"     : _get_str(ws, row, 5),
            })

    return planning


# ------------------------------------------------------------
# LECTURE FEUILLE NUTRITION
# ------------------------------------------------------------
def _read_nutrition(ws) -> dict:
    nutrition = {
        "proteines_min_g_kg"    : _get_float(ws,  4, 2, 1.8),
        "proteines_max_g_kg"    : _get_float(ws,  5, 2, 2.5),
        "lipides_min_g_kg"      : _get_float(ws,  7, 2, 0.8),
        "deficit_entrainement"  : _get_float(ws,  9, 2, 200.0),
        "deficit_repos"         : _get_float(ws, 10, 2, 0.0),
        "calories_min_absolues" : _get_float(ws, 11, 2, 1800.0),
        "repas_par_jour"        : _get_int  (ws, 12, 2, 3),
        "glucides_par_seance"   : {},
    }

    for row in range(15, 21):
        type_seance = _get_str(ws, row, 1)
        if type_seance:
            nutrition["glucides_par_seance"][type_seance] = {
                "min_g_kg"   : _get_float(ws, row, 2, 3.0),
                "max_g_kg"   : _get_float(ws, row, 3, 5.0),
                "moment_cle" : _get_str  (ws, row, 4),
                "notes"      : _get_str  (ws, row, 5),
            }

    return nutrition


# ------------------------------------------------------------
# LECTURE FEUILLE BUDGET
# Inclut maintenant la normalisation du magasin
# ------------------------------------------------------------
def _read_budget(ws) -> dict:
    magasin_raw = _get_str(ws, 8, 2, MAGASIN_FALLBACK)
    magasin     = _normaliser_magasin(magasin_raw)

    return {
        "budget_hebdo_max"     : _get_float(ws,  4, 2, 80.0),
        "budget_quotidien_max" : _get_float(ws,  5, 2, 12.0),
        "budget_repas_max"     : _get_float(ws,  6, 2,  4.0),
        "magasin"              : magasin,        # ✅ Normalisé
        "magasin_raw"          : magasin_raw,    # ✅ Valeur brute Excel
        "priorite_promotions"  : _get_str  (ws,  9, 2, "Oui"),
        "bio_accepte"          : _get_str  (ws, 10, 2, "Non"),
        "marque_distributeur"  : _get_str  (ws, 11, 2, "Oui"),
    }


# ------------------------------------------------------------
# LECTURE FEUILLE ALIMENTS EXCLUS
# ------------------------------------------------------------
def _read_aliments_exclus(ws) -> list:
    aliments_exclus = []

    for row in range(4, ws.max_row + 1):
        nom = _get_str(ws, row, 2)
        if nom:
            aliments_exclus.append({
                "nom"           : nom,
                "type_exclusion": _get_str(ws, row, 3, "contient"),
                "raison"        : _get_str(ws, row, 4),
                "date_ajout"    : _get_str(ws, row, 5),
            })

    return aliments_exclus


# ------------------------------------------------------------
# LECTURE FEUILLE STRUCTURE REPAS
# ------------------------------------------------------------
def _read_structure_repas(ws) -> dict:
    return {
        "semaine" : {
            "petit_dej" : {
                "nb_recettes"   : _get_int(ws,  5, 2, 1),
                "batch_cooking" : _is_oui (ws,  6, 2),
            },
            "dejeuner" : {
                "nb_recettes"   : _get_int(ws,  8, 2, 1),
                "batch_cooking" : _is_oui (ws,  9, 2),
            },
            "diner" : {
                "nb_recettes"   : _get_int(ws, 11, 2, 5),
                "batch_cooking" : _is_oui (ws, 12, 2),
            },
        },
        "weekend" : {
            "petit_dej" : {
                "nb_recettes"   : _get_int(ws,  5, 3, 1),
                "batch_cooking" : _is_oui (ws,  6, 3),
            },
            "dejeuner" : {
                "nb_recettes"   : _get_int(ws,  8, 3, 2),
                "batch_cooking" : _is_oui (ws,  9, 3),
            },
            "diner" : {
                "nb_recettes"   : _get_int(ws, 11, 3, 2),
                "batch_cooking" : _is_oui (ws, 12, 3),
            },
        },
    }


# ------------------------------------------------------------
# LECTURE FEUILLE RECETTES EXCLUES
# ------------------------------------------------------------
def _read_recettes_exclues(ws) -> list:
    recettes_exclues = []

    for row in range(4, ws.max_row + 1):
        nom = _get_str(ws, row, 1)
        if nom:
            recettes_exclues.append({
                "nom"        : nom,
                "raison"     : _get_str(ws, row, 2),
                "date_ajout" : _get_str(ws, row, 3),
            })

    return recettes_exclues


# ------------------------------------------------------------
# VALIDATION DU PROFIL
# ------------------------------------------------------------
def _valider_profil(profil: dict) -> list:
    erreurs = []

    champs_obligatoires = {
        "age"            : "Age",
        "sexe"           : "Sexe",
        "taille_cm"      : "Taille",
        "poids_actuel_kg": "Poids actuel",
        "poids_cible_kg" : "Poids cible",
        "sport"          : "Sport pratiqué",
        "objectif"       : "Objectif",
    }

    for champ, label in champs_obligatoires.items():
        val = profil.get(champ)
        if not val or val == 0:
            erreurs.append(f"[ERREUR] Champ obligatoire manquant : {label}")

    if profil.get("sexe") not in ["M", "F"]:
        erreurs.append("[ERREUR] Sexe doit être 'M' ou 'F'")

    if profil.get("taille_cm", 0) < 100 or profil.get("taille_cm", 0) > 250:
        erreurs.append("[ERREUR] Taille invalide (entre 100 et 250 cm)")

    if profil.get("poids_actuel_kg", 0) < 30 or profil.get("poids_actuel_kg", 0) > 250:
        erreurs.append("[ERREUR] Poids actuel invalide")

    return erreurs


# ------------------------------------------------------------
# FONCTION PRINCIPALE
# ------------------------------------------------------------
def lire_config_athlete() -> dict:
    """
    Lit l'intégralité du fichier athlete_config.xlsm
    et retourne un dictionnaire structuré.

    Le magasin est lu depuis la feuille BUDGET (ligne 8)
    et normalisé automatiquement.
    """
    if not EXCEL_CONFIG_PATH.exists():
        raise FileNotFoundError(
            f"[ERREUR] Fichier de configuration introuvable : {EXCEL_CONFIG_PATH}\n"
            f"[INFO]   Placez athlete_config.xlsm dans optimisation/data/"
        )

    print(f"[...] Lecture configuration : {EXCEL_CONFIG_PATH.name}")

    wb = openpyxl.load_workbook(EXCEL_CONFIG_PATH, data_only=True)

    profil          = _read_profil         (wb["PROFIL"])
    planning        = _read_planning       (wb["PLANNING"])
    nutrition       = _read_nutrition      (wb["NUTRITION"])
    budget          = _read_budget         (wb["BUDGET"])
    aliments_exclus = _read_aliments_exclus(wb["ALIMENTS_EXCLUS"])
    structure_repas = _read_structure_repas(wb["STRUCTURE_REPAS"])
    recettes_exclues = _read_recettes_exclues(wb["RECETTES_EXCLUES"]) \
                       if "RECETTES_EXCLUES" in wb.sheetnames else []

    wb.close()

    # Validation
    erreurs = _valider_profil(profil)
    if erreurs:
        for e in erreurs:
            print(e)
        raise ValueError("[ERREUR] Configuration incomplète.")

    config = {
        "profil"           : profil,
        "planning"         : planning,
        "nutrition"        : nutrition,
        "budget"           : budget,
        "aliments_exclus"  : aliments_exclus,
        "structure_repas"  : structure_repas,
        "recettes_exclues" : recettes_exclues,
    }

    # Récapitulatif
    print(f"[OK] Configuration chargée : {profil.get('nom', 'Athlète')}")
    print(f"     Poids actuel     : {profil['poids_actuel_kg']} kg")
    print(f"     Poids cible      : {profil['poids_cible_kg']} kg")
    print(f"     Sport            : {profil['sport']}")
    print(f"     Magasin          : {budget['magasin'].upper()}")
    print(f"     Budget hebdo     : {budget['budget_hebdo_max']} €")
    print(f"     Aliments exclus  : {len(aliments_exclus)}")
    print(f"     Recettes exclues : {len(recettes_exclues)}")

    return config


# ------------------------------------------------------------
# TEST AUTONOME
# ------------------------------------------------------------
if __name__ == "__main__":
    config = lire_config_athlete()

    print("\n--- PROFIL ---")
    for k, v in config["profil"].items():
        print(f"  {k:<30} : {v}")

    print("\n--- BUDGET & MAGASIN ---")
    for k, v in config["budget"].items():
        print(f"  {k:<30} : {v}")

    print("\n--- PLANNING ---")
    for jour in config["planning"]:
        print(f"  {jour['jour']:<12} {jour['seance']:<35} {jour['intensite']}")

    print("\n--- GLUCIDES PAR SÉANCE ---")
    for seance, val in config["nutrition"]["glucides_par_seance"].items():
        print(f"  {seance:<25} min:{val['min_g_kg']} max:{val['max_g_kg']} g/kg")

    print("\n--- ALIMENTS EXCLUS ---")
    for a in config["aliments_exclus"]:
        print(f"  [{a['type_exclusion']:<10}] {a['nom']}")