# ============================================================
# EXPORT EXCEL DU PLANNING HEBDOMADAIRE
# optimisation/excel/writer.py
#
# Deux modes de fonctionnement :
#
# Mode 1 — Ecriture dans athlete_config.xlsm (production)
#   -> ecrire_resultats_excel()
#   -> Preserve les feuilles de config existantes
#   -> Ecrase uniquement les feuilles de resultats
#   -> Utilise par run_weekly.py (cron automatique)
#
# Mode 2 — Export dans un nouveau fichier (backup/archive)
#   -> exporter_planning()
#   -> Genere un fichier horodate dans output/
#   -> Utilise pour archiver les plannings passes
# ============================================================

import os
import sys
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime
sys.dont_write_bytecode = True


# ------------------------------------------------------------
# CONFIGURATION CHEMINS
# ------------------------------------------------------------
OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "output"
)

FEUILLES_RESULTATS = [
    "PLANNING_SEMAINE",
    "RECETTES",
    "LISTE_COURSES",
    "SUIVI_NUTRITIONNEL",
]

FEUILLES_CONFIG = [
    "PROFIL",
    "PLANNING",
    "NUTRITION",
    "BUDGET",
    "ALIMENTS_EXCLUS",
    "STRUCTURE_REPAS",
]


# ------------------------------------------------------------
# PALETTE DE COULEURS
# ------------------------------------------------------------
COULEURS = {
    "header_bg"  : "1C2B3A",
    "header_fg"  : "FFFFFF",
    "titre_bg"   : "2E4057",
    "titre_fg"   : "FFFFFF",
    "section_bg" : "E8EDF2",
    "section_fg" : "1C2B3A",
    "valeur_bg"  : "FFFFFF",
    "valeur_fg"  : "2C2C2C",
    "accent"     : "C8972B",
    "alt_row"    : "EEF2F5",
    "ok"         : "27AE60",
    "warn"       : "E67E22",
    "erreur"     : "C0392B",
    "petit_dej"  : "2980B9",
    "dejeuner"   : "27AE60",
    "diner"      : "8E44AD",
    "changer"    : "E74C3C",
}


# ------------------------------------------------------------
# STYLES DE BASE
# ------------------------------------------------------------
def _border_thin(color="B0BEC5"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color, bold=False, size=10, italic=False):
    return Font(
        color=hex_color,
        bold=bold,
        size=size,
        italic=italic,
        name="Calibri"
    )


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


def _write_titre(ws, titre, sous_titre, nb_cols=8):
    ws.merge_cells(f"A1:{get_column_letter(nb_cols)}1")
    c           = ws["A1"]
    c.value     = titre.upper()
    c.fill      = _fill(COULEURS["titre_bg"])
    c.font      = _font(COULEURS["titre_fg"], bold=True, size=14)
    c.alignment = _align("left", "center")
    _set_row_height(ws, 1, 32)

    ws.merge_cells(f"A2:{get_column_letter(nb_cols)}2")
    c           = ws["A2"]
    c.value     = sous_titre
    c.fill      = _fill(COULEURS["accent"])
    c.font      = _font("FFFFFF", size=9, italic=True)
    c.alignment = _align("left", "center")
    _set_row_height(ws, 2, 16)


def _write_header_row(ws, row, colonnes, bg=None):
    bg = bg or COULEURS["header_bg"]
    for col, titre in enumerate(colonnes, 1):
        c           = ws.cell(row=row, column=col)
        c.value     = titre
        c.fill      = _fill(bg)
        c.font      = _font(COULEURS["header_fg"], bold=True, size=9)
        c.alignment = _align("center", "center")
        c.border    = _border_thin()
    _set_row_height(ws, row, 20)


def _write_separateur(ws, row, texte, nb_cols=8):
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=nb_cols
    )
    c           = ws.cell(row=row, column=1)
    c.value     = texte.upper()
    c.fill      = _fill(COULEURS["titre_bg"])
    c.font      = _font(COULEURS["accent"], bold=True, size=9)
    c.alignment = _align("left", "center")
    _set_row_height(ws, row, 16)


def _couleur_type_repas(type_repas: str) -> str:
    return {
        "petit_dej" : COULEURS["petit_dej"],
        "dejeuner"  : COULEURS["dejeuner"],
        "diner"     : COULEURS["diner"],
    }.get(type_repas, COULEURS["section_bg"])


# ------------------------------------------------------------
# FEUILLE PLANNING SEMAINE
# Avec bouton CHANGER par recette (col 13)
# Colonnes cachees 14/15/16 pour la macro VBA
# ------------------------------------------------------------
def _build_feuille_planning(ws, resultats: list, config: dict):
    ws.sheet_view.showGridLines = False

    _set_col_width(ws,  1,  5)
    _set_col_width(ws,  2, 20)
    _set_col_width(ws,  3, 14)
    _set_col_width(ws,  4, 38)
    _set_col_width(ws,  5,  8)
    _set_col_width(ws,  6,  7)
    _set_col_width(ws,  7,  7)
    _set_col_width(ws,  8,  7)
    _set_col_width(ws,  9,  8)
    _set_col_width(ws, 10, 10)
    _set_col_width(ws, 11, 14)
    _set_col_width(ws, 12, 32)
    _set_col_width(ws, 13, 12)  # bouton Changer
    _set_col_width(ws, 14,  1)  # nom recette cache
    _set_col_width(ws, 15,  1)  # type repas cache
    _set_col_width(ws, 16,  1)  # jours caches

    profil = config.get("profil", {})
    nom    = profil.get("nom", "Athlete")
    date   = datetime.now().strftime("%d/%m/%Y")

    _write_titre(
        ws,
        f"Planning alimentaire — {nom}",
        f"Semaine du {date} — "
        f"Objectif : {profil.get('objectif', '')} — "
        f"{profil.get('poids_actuel_kg', '')} kg "
        f"-> {profil.get('poids_cible_kg', '')} kg",
        nb_cols=13
    )

    _write_header_row(ws, 3, [
        "N", "Jours", "Type", "Recette",
        "Kcal", "Prot", "Gluc", "Lip",
        "Cout", "Prep", "Batch", "Source",
        "Action"
    ])

    row = 4
    for i, res in enumerate(resultats, 1):
        alt        = (i % 2 == 0)
        bg         = COULEURS["alt_row"] if alt else COULEURS["valeur_bg"]
        jours_str  = " / ".join(res["jours"])
        batch_str  = f"x{len(res['jours'])} portions" if res["batch"] else ""
        mr         = res["macros_reelles"]
        couleur_tr = _couleur_type_repas(res["type_repas"])

        valeurs = [
            (i,                              "center"),
            (jours_str,                      "left"),
            (res["type_repas"],              "center"),
            (res["nom_fr"],                  "left"),
            (f"{mr['calories']:.0f}",        "center"),
            (f"{mr['proteines_g']:.0f}g",    "center"),
            (f"{mr['glucides_g']:.0f}g",     "center"),
            (f"{mr['lipides_g']:.0f}g",      "center"),
            (f"{res['cout']:.2f} e",         "center"),
            (f"{res['temps_prep_min']} min", "center"),
            (batch_str,                      "center"),
            (res.get("source_url", ""),      "left"),
            ("Changer",                      "center"),
        ]

        for col, (val, halign) in enumerate(valeurs, 1):
            c           = ws.cell(row=row, column=col)
            c.value     = val
            c.fill      = _fill(bg)
            c.font      = _font(COULEURS["valeur_fg"], size=9)
            c.alignment = _align(halign, "center")
            c.border    = _border_thin()

        # Style bouton Changer
        ws.cell(row=row, column=13).fill = _fill(COULEURS["changer"])
        ws.cell(row=row, column=13).font = _font("FFFFFF", bold=True, size=9)

        # Colorier cellule type repas
        ws.cell(row=row, column=3).fill = _fill(couleur_tr)
        ws.cell(row=row, column=3).font = _font("FFFFFF", bold=True, size=8)

        # Colonnes cachees pour la macro VBA
        # Col 14 : nom de la recette
        c        = ws.cell(row=row, column=14)
        c.value  = res["nom_fr"]
        c.font   = _font("FFFFFF", size=1)
        c.fill   = _fill("FFFFFF")

        # Col 15 : type de repas
        c        = ws.cell(row=row, column=15)
        c.value  = res["type_repas"]
        c.font   = _font("FFFFFF", size=1)
        c.fill   = _fill("FFFFFF")

        # Col 16 : jours concernes
        c        = ws.cell(row=row, column=16)
        c.value  = jours_str
        c.font   = _font("FFFFFF", size=1)
        c.fill   = _fill("FFFFFF")

        _set_row_height(ws, row, 18)
        row += 1

    # Recapitulatif
    row += 1
    _write_separateur(ws, row, "Recapitulatif semaine", nb_cols=13)
    row += 1

    total_cout = sum(r["cout"] for r in resultats)
    budget_max = config.get("budget", {}).get("budget_hebdo_max", 0)
    nb_ok      = sum(1 for r in resultats if r["statut"] == "ok")

    totaux = [
        ("Generee le",        datetime.now().strftime("%d/%m/%Y a %H:%M")),
        ("Recettes trouvees", f"{nb_ok} / {len(resultats)}"),
        ("Budget utilise",    f"{total_cout:.2f} euros"),
        ("Budget disponible", f"{budget_max:.2f} euros"),
        ("Budget restant",    f"{budget_max - total_cout:.2f} euros"),
    ]

    for label, valeur in totaux:
        c           = ws.cell(row=row, column=1)
        c.value     = label
        c.fill      = _fill(COULEURS["section_bg"])
        c.font      = _font(COULEURS["section_fg"], bold=True, size=9)
        c.border    = _border_thin()

        ws.merge_cells(
            start_row=row, start_column=2,
            end_row=row,   end_column=5
        )
        c           = ws.cell(row=row, column=2)
        c.value     = valeur
        c.fill      = _fill(COULEURS["valeur_bg"])
        c.font      = _font(COULEURS["valeur_fg"], size=9)
        c.border    = _border_thin()
        _set_row_height(ws, row, 18)
        row += 1

    # Note bouton Changer
    row += 1
    ws.merge_cells(f"A{row}:M{row}")
    c           = ws.cell(row=row, column=1)
    c.value     = (
        "Bouton CHANGER : cliquez pour remplacer une recette "
        "qui ne vous convient pas. "
        "Une alternative sera generee automatiquement."
    )
    c.font      = _font(COULEURS["section_fg"], size=8, italic=True)
    c.alignment = _align("left", "center")
    _set_row_height(ws, row, 14)

    ws.sheet_properties.tabColor = COULEURS["titre_bg"]


# ------------------------------------------------------------
# FEUILLE RECETTES DETAILLEES
# ------------------------------------------------------------
def _build_feuille_recettes(ws, resultats: list):
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, 1,  3)
    _set_col_width(ws, 2, 34)
    _set_col_width(ws, 3, 12)
    _set_col_width(ws, 4, 10)
    _set_col_width(ws, 5, 10)
    _set_col_width(ws, 6, 10)
    _set_col_width(ws, 7, 10)
    _set_col_width(ws, 8, 36)

    _write_titre(
        ws,
        "Recettes detaillees",
        "Ingredients adaptes a vos besoins nutritionnels "
        "avec instructions de preparation",
        nb_cols=8
    )

    row = 3

    for res in resultats:
        if res["statut"] != "ok":
            continue

        jours_str  = " / ".join(res["jours"])
        batch_str  = (
            f" [BATCH x{len(res['jours'])} portions]"
            if res["batch"] else ""
        )
        couleur_tr = _couleur_type_repas(res["type_repas"])
        mr         = res["macros_reelles"]
        mc         = res["macros_cibles"]

        # Titre recette
        ws.merge_cells(f"A{row}:H{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = f"{res['nom_fr'].upper()}{batch_str} — {jours_str}"
        c.fill      = _fill(couleur_tr)
        c.font      = _font("FFFFFF", bold=True, size=11)
        c.alignment = _align("left", "center")
        _set_row_height(ws, row, 24)
        row += 1

        # Infos
        infos = "   |   ".join([
            f"Type : {res['type_repas']}",
            f"Seance : {res['seance']}",
            f"Prep : {res['temps_prep_min']} min",
            f"Ratio : x{res['ratio_adaptation']}",
        ])
        ws.merge_cells(f"A{row}:H{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = infos
        c.fill      = _fill(COULEURS["section_bg"])
        c.font      = _font(COULEURS["section_fg"], size=8, italic=True)
        c.alignment = _align("left", "center")
        _set_row_height(ws, row, 14)
        row += 1

        # Macros
        ws.merge_cells(f"A{row}:H{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = (
            f"Reelles : {mr['calories']:.0f} kcal"
            f" P:{mr['proteines_g']:.0f}g"
            f" G:{mr['glucides_g']:.0f}g"
            f" L:{mr['lipides_g']:.0f}g"
            f"     |     "
            f"Cibles : {mc['calories']:.0f} kcal"
            f" P:{mc['proteines_g']:.0f}g"
            f" G:{mc['glucides_g']:.0f}g"
            f" L:{mc['lipides_g']:.0f}g"
        )
        c.fill      = _fill(COULEURS["valeur_bg"])
        c.font      = _font(COULEURS["valeur_fg"], size=8)
        c.alignment = _align("left", "center")
        _set_row_height(ws, row, 14)
        row += 1

        # Ingredients
        if res["ingredients"]:
            _write_header_row(ws, row, [
                "", "Ingredient", "Quantite (g)",
                "Kcal", "Prot", "Gluc", "Lip", ""
            ], bg=COULEURS["section_fg"])
            row += 1

            for j, ing in enumerate(res["ingredients"]):
                bg = COULEURS["alt_row"] if j % 2 == 0 else COULEURS["valeur_bg"]
                valeurs = [
                    ("",                         "center"),
                    (ing["nom_fr"],              "left"),
                    (f"{ing['quantite_g']:.0f}", "center"),
                    (f"{ing['calories']:.0f}",   "center"),
                    (f"{ing['proteines_g']:.1f}","center"),
                    (f"{ing['glucides_g']:.1f}", "center"),
                    (f"{ing['lipides_g']:.1f}",  "center"),
                    ("",                         "center"),
                ]
                for col, (val, halign) in enumerate(valeurs, 1):
                    c           = ws.cell(row=row, column=col)
                    c.value     = val
                    c.fill      = _fill(bg)
                    c.font      = _font(COULEURS["valeur_fg"], size=9)
                    c.alignment = _align(halign, "center")
                    c.border    = _border_thin()
                _set_row_height(ws, row, 16)
                row += 1

        # Instructions
        if res["instructions"]:
            row += 1
            ws.merge_cells(f"A{row}:H{row}")
            c           = ws.cell(row=row, column=1)
            c.value     = "PREPARATION"
            c.fill      = _fill(COULEURS["titre_bg"])
            c.font      = _font(COULEURS["accent"], bold=True, size=9)
            c.alignment = _align("left", "center")
            _set_row_height(ws, row, 16)
            row += 1

            for step in res["instructions"]:
                ws.merge_cells(f"A{row}:H{row}")
                c           = ws.cell(row=row, column=1)
                c.value     = f"  {step['etape']}. {step['instruction_fr']}"
                c.fill      = _fill(COULEURS["valeur_bg"])
                c.font      = _font(COULEURS["valeur_fg"], size=9)
                c.alignment = _align("left", "center", wrap=True)
                _set_row_height(ws, row, 20)
                row += 1

        # Source
        if res.get("source_url"):
            ws.merge_cells(f"A{row}:H{row}")
            c           = ws.cell(row=row, column=1)
            c.value     = f"Source : {res['source_url']}"
            c.font      = _font(COULEURS["section_fg"], size=8, italic=True)
            c.alignment = _align("left", "center")
            _set_row_height(ws, row, 14)
            row += 1

        row += 2

    ws.sheet_properties.tabColor = COULEURS["accent"]


# ------------------------------------------------------------
# FEUILLE LISTE DE COURSES
# ------------------------------------------------------------
def _build_feuille_courses(ws, resultats: list):
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, 1,  4)
    _set_col_width(ws, 2, 36)
    _set_col_width(ws, 3, 18)
    _set_col_width(ws, 4, 12)
    _set_col_width(ws, 5, 30)

    _write_titre(
        ws,
        "Liste de courses hebdomadaire",
        "Ingredients consolides pour la semaine — "
        "a adapter selon les promotions disponibles",
        nb_cols=5
    )

    courses = {}
    for res in resultats:
        if res["statut"] != "ok":
            continue
        nb_portions = len(res["jours"]) if res["batch"] else 1
        for ing in res["ingredients"]:
            nom = ing["nom_fr"].strip().lower()
            if nom not in courses:
                courses[nom] = {
                    "nom"        : ing["nom_fr"],
                    "quantite_g" : 0.0,
                    "repas"      : [],
                }
            courses[nom]["quantite_g"] += ing["quantite_g"] * nb_portions
            if res["nom_fr"] not in courses[nom]["repas"]:
                courses[nom]["repas"].append(res["nom_fr"])

    courses_tries = sorted(
        courses.values(),
        key=lambda x: x["nom"].lower()
    )

    _write_header_row(ws, 3, [
        "N", "Ingredient", "Quantite totale", "Achete", "Utilise dans"
    ])

    row = 4
    for i, ing in enumerate(courses_tries, 1):
        bg      = COULEURS["alt_row"] if i % 2 == 0 else COULEURS["valeur_bg"]
        qte_kg  = ing["quantite_g"] / 1000
        qte_str = (
            f"{ing['quantite_g']:.0f} g"
            if qte_kg < 1
            else f"{qte_kg:.2f} kg"
        )

        valeurs = [
            (i,                           "center"),
            (ing["nom"],                  "left"),
            (qte_str,                     "center"),
            ("",                          "center"),
            (", ".join(ing["repas"][:3]), "left"),
        ]

        for col, (val, halign) in enumerate(valeurs, 1):
            c           = ws.cell(row=row, column=col)
            c.value     = val
            c.fill      = _fill(bg)
            c.font      = _font(COULEURS["valeur_fg"], size=9)
            c.alignment = _align(halign, "center")
            c.border    = _border_thin()

        _set_row_height(ws, row, 18)
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    c           = ws.cell(row=row, column=1)
    c.value     = (
        f"Total : {len(courses_tries)} ingredients — "
        f"Generee le {datetime.now().strftime('%d/%m/%Y a %H:%M')}"
    )
    c.font      = _font(COULEURS["section_fg"], size=8, italic=True)
    c.alignment = _align("left", "center")
    _set_row_height(ws, row, 14)

    ws.sheet_properties.tabColor = "27AE60"


# ------------------------------------------------------------
# FEUILLE SUIVI NUTRITIONNEL
# ------------------------------------------------------------
def _build_feuille_nutrition(ws, resultats: list, planning_semaine: list):
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, 1, 16)
    _set_col_width(ws, 2, 30)
    _set_col_width(ws, 3,  9)
    _set_col_width(ws, 4,  9)
    _set_col_width(ws, 5,  9)
    _set_col_width(ws, 6,  9)
    _set_col_width(ws, 7,  9)
    _set_col_width(ws, 8,  9)

    _write_titre(
        ws,
        "Suivi nutritionnel hebdomadaire",
        "Comparaison macros cibles vs macros reelles — "
        "Vert = ecart < 10% / Orange = ecart < 20% / Rouge = ecart > 20%",
        nb_cols=8
    )

    _write_header_row(ws, 3, [
        "Jour", "Recette",
        "Kcal cible", "Kcal reel",
        "Prot cible", "Prot reel",
        "Gluc cible", "Gluc reel",
    ])

    row = 4
    for res in resultats:
        jours_str = " / ".join(res["jours"])
        mr        = res["macros_reelles"]
        mc        = res["macros_cibles"]
        bg        = COULEURS["alt_row"] if row % 2 == 0 else COULEURS["valeur_bg"]

        ecart_pct = (
            abs(mr["calories"] - mc["calories"])
            / max(mc["calories"], 1) * 100
        )
        if ecart_pct <= 10:
            couleur_ecart = COULEURS["ok"]
        elif ecart_pct <= 20:
            couleur_ecart = COULEURS["warn"]
        else:
            couleur_ecart = COULEURS["erreur"]

        valeurs = [
            (jours_str,                   "left",   bg),
            (res["nom_fr"],               "left",   bg),
            (f"{mc['calories']:.0f}",     "center", bg),
            (f"{mr['calories']:.0f}",     "center", couleur_ecart),
            (f"{mc['proteines_g']:.0f}g", "center", bg),
            (f"{mr['proteines_g']:.0f}g", "center", bg),
            (f"{mc['glucides_g']:.0f}g",  "center", bg),
            (f"{mr['glucides_g']:.0f}g",  "center", bg),
        ]

        for col, (val, halign, bg_cell) in enumerate(valeurs, 1):
            c           = ws.cell(row=row, column=col)
            c.value     = val
            c.fill      = _fill(bg_cell)
            c.font      = _font(
                "FFFFFF"
                if bg_cell not in [COULEURS["alt_row"], COULEURS["valeur_bg"]]
                else COULEURS["valeur_fg"],
                size=9
            )
            c.alignment = _align(halign, "center")
            c.border    = _border_thin()

        _set_row_height(ws, row, 18)
        row += 1

    row += 1
    _write_separateur(ws, row, "Moyennes sur la semaine", nb_cols=8)
    row += 1

    nb = len(resultats)
    if nb > 0:
        moyennes = [
            ("Moyenne calories/repas",
             f"{sum(r['macros_reelles']['calories']    for r in resultats)/nb:.0f} kcal"),
            ("Moyenne proteines/repas",
             f"{sum(r['macros_reelles']['proteines_g'] for r in resultats)/nb:.0f} g"),
            ("Moyenne glucides/repas",
             f"{sum(r['macros_reelles']['glucides_g']  for r in resultats)/nb:.0f} g"),
            ("Moyenne lipides/repas",
             f"{sum(r['macros_reelles']['lipides_g']   for r in resultats)/nb:.0f} g"),
            ("Total calories semaine",
             f"{sum(r['macros_reelles']['calories']    for r in resultats):.0f} kcal"),
        ]

        for label, valeur in moyennes:
            c           = ws.cell(row=row, column=1)
            c.value     = label
            c.fill      = _fill(COULEURS["section_bg"])
            c.font      = _font(COULEURS["section_fg"], bold=True, size=9)
            c.border    = _border_thin()

            ws.merge_cells(
                start_row=row, start_column=2,
                end_row=row,   end_column=4
            )
            c           = ws.cell(row=row, column=2)
            c.value     = valeur
            c.fill      = _fill(COULEURS["valeur_bg"])
            c.font      = _font(COULEURS["valeur_fg"], size=9)
            c.border    = _border_thin()
            _set_row_height(ws, row, 18)
            row += 1

    ws.sheet_properties.tabColor = "8E44AD"

# ------------------------------------------------------------
# FEUILLE RECETTES EXCLUES
# Creee une seule fois, jamais ecrasee
# Remplie automatiquement par la macro VBA
# quand l utilisateur clique sur Changer
# ------------------------------------------------------------
def _build_feuille_recettes_exclues(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COULEURS["changer"]

    _set_col_width(ws, 1, 40)
    _set_col_width(ws, 2, 25)
    _set_col_width(ws, 3, 15)

    _write_titre(
        ws,
        "Recettes exclues",
        "Recettes ajoutees via le bouton Changer. "
        "Ne seront jamais reproposees.",
        nb_cols=3
    )

    _write_header_row(ws, 3, [
        "Nom de la recette",
        "Raison",
        "Date"
    ])

    # Ligne exemple
    exemples = [
        ("Exemple : Chicken Brown Rice", "N aime pas", datetime.now().strftime("%d/%m/%Y")),
    ]
    for i, (nom, raison, date) in enumerate(exemples, start=4):
        bg = COULEURS["alt_row"] if i % 2 == 0 else COULEURS["valeur_bg"]
        for col, val in enumerate([nom, raison, date], 1):
            c           = ws.cell(row=i, column=col)
            c.value     = val
            c.fill      = _fill(bg)
            c.font      = _font(COULEURS["section_fg"], size=9, italic=True)
            c.alignment = _align("left", "center")
            c.border    = _border_thin()
        _set_row_height(ws, i, 18)


# ------------------------------------------------------------
# ECRITURE DANS ATHLETE_CONFIG.XLSM (MODE PRODUCTION)
# ------------------------------------------------------------
def ecrire_resultats_excel(
    resultats        : list,
    config           : dict,
    planning_semaine : list,
) -> str:

    sys.path.insert(0, os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "..")
    ))
    from optimisation.config import EXCEL_CONFIG_PATH

    # Verifier que le fichier n est pas ouvert
    try:
        with open(EXCEL_CONFIG_PATH, "a"):
            pass
    except PermissionError:
        raise PermissionError(
            f"\n[ERREUR] Fermez athlete_config.xlsm dans Excel "
            f"avant de lancer le script.\n"
            f"         {EXCEL_CONFIG_PATH}"
        )

    print(f"\n[...] Ecriture des resultats dans : {EXCEL_CONFIG_PATH}")

    wb = openpyxl.load_workbook(
        EXCEL_CONFIG_PATH,
        keep_vba=True
    )

    # Supprimer uniquement les feuilles de resultats
    for nom in FEUILLES_RESULTATS:
        if nom in wb.sheetnames:
            del wb[nom]
            print(f"      Feuille '{nom}' supprimee et recreee")

    # Creer RECETTES_EXCLUES uniquement si elle n existe pas
    # On ne la supprime jamais car elle contient les choix utilisateur
    if "RECETTES_EXCLUES" not in wb.sheetnames:
        ws_exclues = wb.create_sheet("RECETTES_EXCLUES")
        _build_feuille_recettes_exclues(ws_exclues)
        print(f"      Feuille 'RECETTES_EXCLUES' creee")

    # Creer les nouvelles feuilles de resultats
    ws_planning  = wb.create_sheet("PLANNING_SEMAINE")
    ws_recettes  = wb.create_sheet("RECETTES")
    ws_courses   = wb.create_sheet("LISTE_COURSES")
    ws_nutrition = wb.create_sheet("SUIVI_NUTRITIONNEL")

    _build_feuille_planning (ws_planning,  resultats, config)
    _build_feuille_recettes (ws_recettes,  resultats)
    _build_feuille_courses  (ws_courses,   resultats)
    _build_feuille_nutrition(ws_nutrition, resultats, planning_semaine)

    wb.save(EXCEL_CONFIG_PATH)
    print(f"[OK] Resultats ecrits dans : {EXCEL_CONFIG_PATH}")
    print(f"     Feuilles config preservees  : {FEUILLES_CONFIG}")
    print(f"     Feuilles resultats ecrites  : {FEUILLES_RESULTATS}")
    print(f"     Feuille RECETTES_EXCLUES    : preservee")

    return EXCEL_CONFIG_PATH

def _mettre_a_jour_repas_excel(
    nom_recette_exclue : str,
    nouveau_repas      : dict,
    type_repas         : str,
    jours              : list,
) -> None:
    """
    Met a jour une ligne specifique dans PLANNING_SEMAINE
    et met a jour la feuille RECETTES avec la nouvelle recette.
    """
    sys.path.insert(0, os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "..")
    ))
    from optimisation.config import EXCEL_CONFIG_PATH

    # Verifier que le fichier n est pas ouvert
    try:
        with open(EXCEL_CONFIG_PATH, "a"):
            pass
    except PermissionError:
        raise PermissionError(
            f"[ERREUR] Fermez athlete_config.xlsm avant de regenerer."
        )

    wb              = openpyxl.load_workbook(EXCEL_CONFIG_PATH, keep_vba=True)
    ws_planning     = wb["PLANNING_SEMAINE"]
    mr              = nouveau_repas["macros_reelles"]
    jours_str       = " / ".join(jours)

    # Chercher la ligne a mettre a jour
    for row in range(4, ws_planning.max_row + 1):
        nom_cellule = ws_planning.cell(row=row, column=14).value
        if nom_cellule and nom_cellule.lower() == nom_recette_exclue.lower():

            # Mettre a jour le nom de la recette
            ws_planning.cell(row=row, column=4).value  = nouveau_repas["nom_fr"]
            ws_planning.cell(row=row, column=5).value  = f"{mr['calories']:.0f}"
            ws_planning.cell(row=row, column=6).value  = f"{mr['proteines_g']:.0f}g"
            ws_planning.cell(row=row, column=7).value  = f"{mr['glucides_g']:.0f}g"
            ws_planning.cell(row=row, column=8).value  = f"{mr['lipides_g']:.0f}g"
            ws_planning.cell(row=row, column=9).value  = f"{nouveau_repas['cout']:.2f} e"
            ws_planning.cell(row=row, column=10).value = f"{nouveau_repas['temps_prep_min']} min"
            ws_planning.cell(row=row, column=12).value = nouveau_repas.get("source_url", "")

            # Mettre a jour les colonnes cachees
            ws_planning.cell(row=row, column=14).value = nouveau_repas["nom_fr"]

            print(f"[OK] Ligne {row} mise a jour : {nouveau_repas['nom_fr']}")
            break

    # Mettre a jour la feuille RECETTES
    if "RECETTES" in wb.sheetnames:
        ws_recettes = wb["RECETTES"]

        # Trouver et remplacer le bloc de la recette exclue
        for row in range(3, ws_recettes.max_row + 1):
            val = ws_recettes.cell(row=row, column=1).value
            if val and nom_recette_exclue.upper() in str(val).upper():
                # Remplacer le titre
                ws_recettes.cell(row=row, column=1).value = (
                    f"{nouveau_repas['nom_fr'].upper()} — {jours_str}"
                )
                print(f"[OK] Feuille RECETTES mise a jour ligne {row}")
                break

    wb.save(EXCEL_CONFIG_PATH)
    print(f"[OK] Excel mis a jour : {EXCEL_CONFIG_PATH}")

# ------------------------------------------------------------
# EXPORT DANS UN NOUVEAU FICHIER (MODE ARCHIVE)
# ------------------------------------------------------------
def exporter_planning(
    resultats        : list,
    config           : dict,
    planning_semaine : list,
    nom_fichier      : str = None,
) -> str:

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if nom_fichier is None:
        date_str    = datetime.now().strftime("%Y%m%d_%H%M")
        nom_athlete = (
            config.get("profil", {})
            .get("nom", "athlete")
            .replace(" ", "_")
        )
        nom_fichier = f"planning_{nom_athlete}_{date_str}.xlsx"

    output_path = os.path.join(OUTPUT_DIR, nom_fichier)

    print(f"\n[...] Export archive : {output_path}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_planning  = wb.create_sheet("PLANNING_SEMAINE")
    ws_recettes  = wb.create_sheet("RECETTES")
    ws_courses   = wb.create_sheet("LISTE_COURSES")
    ws_nutrition = wb.create_sheet("SUIVI_NUTRITIONNEL")

    _build_feuille_planning (ws_planning,  resultats, config)
    _build_feuille_recettes (ws_recettes,  resultats)
    _build_feuille_courses  (ws_courses,   resultats)
    _build_feuille_nutrition(ws_nutrition, resultats, planning_semaine)

    wb.save(output_path)
    print(f"[OK] Archive generee : {output_path}")

    return output_path


# ------------------------------------------------------------
# TEST AUTONOME
# ------------------------------------------------------------
if __name__ == "__main__":
    sys.path.insert(0, os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "..")
    ))

    from optimisation.excel.reader            import lire_config_athlete
    from optimisation.engine.calories         import calcul_calories_semaine
    from optimisation.engine.macros           import calcul_macros_semaine
    from optimisation.planning.semaine        import construire_planning_semaine
    from optimisation.engine.recipe_optimizer import generer_semaine_recettes

    config             = lire_config_athlete()
    planning_calorique = calcul_calories_semaine(config)
    planning_macros    = calcul_macros_semaine(config, planning_calorique)
    planning_semaine   = construire_planning_semaine(
        config, planning_calorique, planning_macros
    )
    resultats = generer_semaine_recettes(config, planning_semaine)

    # Mode 1 : ecrire dans athlete_config.xlsm
    ecrire_resultats_excel(resultats, config, planning_semaine)

    # Mode 2 : exporter une archive horodatee
    chemin = exporter_planning(resultats, config, planning_semaine)
    print(f"\n[OK] Export termine : {chemin}")
