# ============================================================
# SUIVI DU POIDS HEBDOMADAIRE
# suivi/tracking/poids.py
#
# Lit l historique du poids depuis athlete_config.xlsx
# Calcule les tendances et projections
# Determine si l objectif est en bonne voie
#
# Sources scientifiques :
#   - Hall et al. 2012 — modele de perte de poids
#     The Lancet
#   - Garthe et al. 2011 — taux de perte optimal athlete
#     International Journal of Sport Nutrition
#   - Helms et al. 2014 — suivi composition corporelle
#     Journal ISSN
# ============================================================

import os
import sys
import openpyxl
from datetime import datetime, timedelta
sys.dont_write_bytecode = True

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import (
    EXCEL_CONFIG_PATH,
    get_parametres_objectif,
)


# ------------------------------------------------------------
# LECTURE DE L HISTORIQUE POIDS DEPUIS EXCEL
# ------------------------------------------------------------
def lire_historique_poids() -> list:
    """
    Lit la feuille SUIVI_POIDS dans athlete_config.xlsx.

    Format attendu de la feuille SUIVI_POIDS :
      Ligne 1 : Titre
      Ligne 2 : Sous-titre
      Ligne 3 : Headers
        Col A : Date (format dd/mm/yyyy)
        Col B : Poids (kg)
        Col C : Notes (optionnel)
      Lignes 4+ : Donnees

    Retourne une liste triee par date croissante :
    [
        {
            "date"   : datetime,
            "poids"  : float,
            "notes"  : str,
            "semaine": int,
            "annee"  : int,
        },
        ...
    ]
    """
    if not os.path.exists(EXCEL_CONFIG_PATH):
        raise FileNotFoundError(
            f"[ERREUR] Fichier config introuvable : {EXCEL_CONFIG_PATH}"
        )

    wb = openpyxl.load_workbook(EXCEL_CONFIG_PATH, data_only=True)

    if "SUIVI_POIDS" not in wb.sheetnames:
        wb.close()
        print("[WARN] Feuille 'SUIVI_POIDS' introuvable dans athlete_config.xlsx")
        print("[INFO] Creez la feuille via : python suivi/tracking/create_suivi_poids.py")
        return []

    ws      = wb["SUIVI_POIDS"]
    entrees = []
    row     = 4

    while ws.cell(row=row, column=1).value:
        date_val  = ws.cell(row=row, column=1).value
        poids_val = ws.cell(row=row, column=2).value
        notes_val = ws.cell(row=row, column=3).value or ""

        if date_val and poids_val:
            # Gerer les differents formats de date
            if isinstance(date_val, datetime):
                date_obj = date_val
            elif isinstance(date_val, str):
                try:
                    date_obj = datetime.strptime(date_val, "%d/%m/%Y")
                except ValueError:
                    try:
                        date_obj = datetime.strptime(date_val, "%Y-%m-%d")
                    except ValueError:
                        print(f"[WARN] Format de date invalide ligne {row} : {date_val}")
                        row += 1
                        continue
            else:
                row += 1
                continue

            try:
                poids = float(poids_val)
                entrees.append({
                    "date"   : date_obj,
                    "poids"  : poids,
                    "notes"  : str(notes_val).strip(),
                    "semaine": date_obj.isocalendar()[1],
                    "annee"  : date_obj.year,
                })
            except (ValueError, TypeError):
                print(f"[WARN] Poids invalide ligne {row} : {poids_val}")

        row += 1

    wb.close()

    # Trier par date croissante
    entrees.sort(key=lambda x: x["date"])

    print(f"[OK] {len(entrees)} releves de poids charges")
    return entrees


# ------------------------------------------------------------
# CALCUL DES TENDANCES
# ------------------------------------------------------------
def calculer_tendances(historique: list) -> dict:
    """
    Calcule les tendances de poids a partir de l historique.

    Retourne :
    {
        "poids_actuel"          : dernier poids releve,
        "poids_initial"         : premier poids releve,
        "variation_totale_kg"   : poids actuel - poids initial,
        "nb_semaines"           : duree totale en semaines,
        "variation_moyenne_semaine" : variation moyenne par semaine,
        "variation_derniere_semaine": variation sur la derniere semaine,
        "variation_2_dernieres_semaines": variation sur 2 semaines,
        "tendance"              : "perte" / "gain" / "stable",
        "poids_par_semaine"     : liste [(semaine, poids), ...],
    }
    """
    if len(historique) < 2:
        return {
            "poids_actuel"                  : historique[-1]["poids"] if historique else 0,
            "poids_initial"                 : historique[0]["poids"]  if historique else 0,
            "variation_totale_kg"           : 0.0,
            "nb_semaines"                   : 0,
            "variation_moyenne_semaine"     : 0.0,
            "variation_derniere_semaine"    : 0.0,
            "variation_2_dernieres_semaines": 0.0,
            "tendance"                      : "stable",
            "poids_par_semaine"             : [],
        }

    poids_actuel  = historique[-1]["poids"]
    poids_initial = historique[0]["poids"]
    date_debut    = historique[0]["date"]
    date_fin      = historique[-1]["date"]

    # Nombre de semaines
    delta_jours = (date_fin - date_debut).days
    nb_semaines = max(delta_jours / 7, 1)

    # Variation totale
    variation_totale = round(poids_actuel - poids_initial, 2)

    # Variation moyenne par semaine
    variation_moy = round(variation_totale / nb_semaines, 3)

    # Variation derniere semaine
    if len(historique) >= 2:
        variation_derniere = round(
            historique[-1]["poids"] - historique[-2]["poids"], 2
        )
    else:
        variation_derniere = 0.0

    # Variation 2 dernieres semaines
    if len(historique) >= 3:
        variation_2sem = round(
            historique[-1]["poids"] - historique[-3]["poids"], 2
        )
    else:
        variation_2sem = variation_derniere

    # Tendance
    if abs(variation_moy) < 0.05:
        tendance = "stable"
    elif variation_moy < 0:
        tendance = "perte"
    else:
        tendance = "gain"

    # Poids par semaine pour graphique
    poids_par_semaine = [
        {
            "date"   : h["date"].strftime("%d/%m/%Y"),
            "poids"  : h["poids"],
            "semaine": h["semaine"],
        }
        for h in historique
    ]

    return {
        "poids_actuel"                  : poids_actuel,
        "poids_initial"                 : poids_initial,
        "variation_totale_kg"           : variation_totale,
        "nb_semaines"                   : round(nb_semaines, 1),
        "variation_moyenne_semaine"     : variation_moy,
        "variation_derniere_semaine"    : variation_derniere,
        "variation_2_dernieres_semaines": variation_2sem,
        "tendance"                      : tendance,
        "poids_par_semaine"             : poids_par_semaine,
    }


# ------------------------------------------------------------
# PROJECTION VERS L OBJECTIF
# ------------------------------------------------------------
def calculer_projection(
    tendances    : dict,
    poids_cible  : float,
    objectif     : str,
) -> dict:
    """
    Projette la date d atteinte de l objectif
    en fonction de la tendance actuelle.

    Source : Hall et al. 2012
    "Quantification de l effet du desequilibre energetique
     sur le poids corporel" — The Lancet
    "La projection lineaire est une approximation valide
     sur des periodes de 8-12 semaines"

    Retourne :
    {
        "poids_cible"            : objectif en kg,
        "ecart_restant_kg"       : poids_actuel - poids_cible,
        "semaines_restantes"     : projection nombre de semaines,
        "date_objectif_estimee"  : date estimee d atteinte,
        "sur_la_bonne_voie"      : bool,
        "message"                : message d analyse,
    }
    """
    params          = get_parametres_objectif(objectif)
    direction       = params["direction"]
    poids_actuel    = tendances["poids_actuel"]
    variation_moy   = tendances["variation_moyenne_semaine"]

    ecart_restant = round(poids_actuel - poids_cible, 2)

    # Verifier si l objectif est deja atteint
    if abs(ecart_restant) < 0.5:
        return {
            "poids_cible"           : poids_cible,
            "ecart_restant_kg"      : ecart_restant,
            "semaines_restantes"    : 0,
            "date_objectif_estimee" : datetime.now().strftime("%d/%m/%Y"),
            "sur_la_bonne_voie"     : True,
            "message"               : "Objectif atteint ou tres proche.",
        }

    # Projection selon la tendance actuelle
    if abs(variation_moy) < 0.05:
        semaines_restantes    = None
        date_objectif_estimee = None
        sur_la_bonne_voie     = False
        message               = (
            "Poids stable depuis plusieurs semaines. "
            "Ajustement nutritionnel recommande."
        )
    else:
        semaines_restantes = abs(ecart_restant / variation_moy)
        semaines_restantes = round(semaines_restantes, 1)

        date_objectif = datetime.now() + timedelta(
            weeks=semaines_restantes
        )
        date_objectif_estimee = date_objectif.strftime("%d/%m/%Y")

        # Determiner si on est sur la bonne voie
        if direction == "perte":
            sur_la_bonne_voie = (
                variation_moy < 0 and
                abs(variation_moy) >= params["perte_optimale_min_kg"]
            )
        elif direction == "gain":
            sur_la_bonne_voie = (
                variation_moy > 0 and
                variation_moy >= params["perte_optimale_min_kg"]
            )
        else:
            sur_la_bonne_voie = abs(variation_moy) < 0.2

        if sur_la_bonne_voie:
            message = (
                f"Progression optimale. "
                f"Objectif estime dans {semaines_restantes:.0f} semaines "
                f"({date_objectif_estimee})."
            )
        else:
            message = (
                f"Progression hors de la fourchette optimale. "
                f"Ajustement recommande. "
                f"Projection actuelle : {semaines_restantes:.0f} semaines."
            )

    return {
        "poids_cible"           : poids_cible,
        "ecart_restant_kg"      : ecart_restant,
        "semaines_restantes"    : semaines_restantes,
        "date_objectif_estimee" : date_objectif_estimee,
        "sur_la_bonne_voie"     : sur_la_bonne_voie,
        "message"               : message,
    }


# ------------------------------------------------------------
# ANALYSE COMPLETE DU SUIVI POIDS
# ------------------------------------------------------------
def analyser_poids(config: dict) -> dict:
    """
    Fonction principale d analyse du suivi poids.

    Lit l historique, calcule les tendances et la projection.

    Retourne un dictionnaire complet utilisable par
    adaptation.py pour ajuster les parametres nutritionnels.
    """
    profil       = config["profil"]
    poids_cible  = profil["poids_cible_kg"]
    objectif     = profil["objectif"]
    params       = get_parametres_objectif(objectif)

    # Lire l historique
    historique = lire_historique_poids()

    if not historique:
        return {
            "statut"     : "pas_de_donnees",
            "message"    : "Aucun releve de poids disponible.",
            "historique" : [],
            "tendances"  : {},
            "projection" : {},
            "params"     : params,
        }

    # Calculer les tendances
    tendances = calculer_tendances(historique)

    # Calculer la projection
    projection = calculer_projection(tendances, poids_cible, objectif)

    # Statut global
    var_sem = tendances["variation_derniere_semaine"]
    var_2sem = tendances["variation_2_dernieres_semaines"]

    if params["direction"] == "perte":
        if var_sem < -params["seuil_changement_rapide_kg"]:
            statut = "perte_trop_rapide"
        elif abs(var_2sem) < params["seuil_stagnation_kg"] * 2:
            statut = "stagnation"
        elif (params["perte_optimale_min_kg"] <=
              abs(var_sem) <=
              params["perte_optimale_max_kg"]):
            statut = "optimal"
        else:
            statut = "sous_optimal"

    elif params["direction"] == "gain":
        if var_sem > params["seuil_changement_rapide_kg"]:
            statut = "gain_trop_rapide"
        elif var_sem < params["seuil_stagnation_kg"]:
            statut = "stagnation"
        elif (params["perte_optimale_min_kg"] <=
              var_sem <=
              params["perte_optimale_max_kg"]):
            statut = "optimal"
        else:
            statut = "sous_optimal"

    else:
        # Maintien ou performance
        if abs(var_sem) > params["seuil_changement_rapide_kg"]:
            statut = "derive"
        else:
            statut = "stable"

    return {
        "statut"          : statut,
        "message"         : projection["message"],
        "historique"      : historique,
        "tendances"       : tendances,
        "projection"      : projection,
        "params"          : params,
        "objectif_detecte": params.get("objectif_detecte", objectif),
    }


# ------------------------------------------------------------
# AFFICHAGE
# ------------------------------------------------------------
def afficher_analyse_poids(analyse: dict) -> None:

    tendances  = analyse.get("tendances",  {})
    projection = analyse.get("projection", {})

    print("\n" + "=" * 65)
    print("  SUIVI POIDS HEBDOMADAIRE")
    print("=" * 65)

    if analyse["statut"] == "pas_de_donnees":
        print("  Aucun releve de poids disponible.")
        print("  Renseignez la feuille SUIVI_POIDS dans athlete_config.xlsx")
        print("=" * 65)
        return

    print(f"  Objectif detecte  : {analyse['objectif_detecte'].upper()}")
    print(f"  Statut            : {analyse['statut'].upper()}")
    print()

    print(f"  Poids initial     : {tendances.get('poids_initial', 0):.1f} kg")
    print(f"  Poids actuel      : {tendances.get('poids_actuel', 0):.1f} kg")
    print(f"  Poids cible       : {projection.get('poids_cible', 0):.1f} kg")
    print(f"  Ecart restant     : {projection.get('ecart_restant_kg', 0):+.1f} kg")
    print()

    print(f"  Variation totale  : {tendances.get('variation_totale_kg', 0):+.2f} kg")
    print(f"  Variation moy/sem : {tendances.get('variation_moyenne_semaine', 0):+.3f} kg")
    print(f"  Variation S-1     : {tendances.get('variation_derniere_semaine', 0):+.2f} kg")
    print(f"  Variation S-2     : {tendances.get('variation_2_dernieres_semaines', 0):+.2f} kg")
    print()

    if projection.get("semaines_restantes") is not None:
        print(f"  Semaines restantes      : {projection['semaines_restantes']:.0f}")
        print(f"  Date objectif estimee   : {projection.get('date_objectif_estimee', 'N/A')}")
        print(f"  Sur la bonne voie       : {'Oui' if projection.get('sur_la_bonne_voie') else 'Non'}")

    print()
    print(f"  Analyse : {analyse['message']}")
    print("=" * 65)


# ------------------------------------------------------------
# TEST AUTONOME
# ------------------------------------------------------------
if __name__ == "__main__":
    sys.path.insert(0, os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "..", "..")
    ))

    from optimisation.excel.reader import lire_config_athlete

    config  = lire_config_athlete()
    analyse = analyser_poids(config)
    afficher_analyse_poids(analyse)