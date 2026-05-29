# ============================================================
# RECALIBRAGE AUTOMATIQUE DES PARAMETRES NUTRITIONNELS
# suivi/tracking/adaptation.py
#
# Analyse le statut de poids hebdomadaire et ajuste
# automatiquement les parametres dans athlete_config.xlsm :
#   - Deficit ou surplus calorique
#   - Glucides selon le type de seance
#
# Sources scientifiques :
#   - Hall et al. 2012 — adaptation metabolique
#     The Lancet
#   - Helms et al. 2014 — ajustements deficit athlete
#     Journal ISSN
#   - Burke et al. 2011 — periodisation glucidique
#     Journal of Sports Sciences
#   - Barakat et al. 2020 — recomposition corporelle
#     Strength and Conditioning Journal
#   - Antonio et al. 2020 — prise de masse ISSN
#   - Loucks et al. 2011 — energie disponible minimum
#     Journal of Sports Sciences
# ============================================================

import os
import sys
import openpyxl
from datetime import datetime
sys.dont_write_bytecode = True

sys.path.insert(0, str(Path(__file__).parent.parent.parent.resolve()))
from config import EXCEL_CONFIG_PATH, get_parametres_objectif


# ------------------------------------------------------------
# LECTURE DES PARAMETRES NUTRITIONNELS ACTUELS
# ------------------------------------------------------------
def _lire_parametres_actuels(ws_nutrition) -> dict:
    """
    Lit les parametres nutritionnels actuels
    depuis la feuille NUTRITION de athlete_config.xlsm.

    Structure attendue (identique a reader.py) :
      Ligne 4  : Proteines minimum
      Ligne 5  : Proteines maximum
      Ligne 7  : Lipides minimum
      Ligne 9  : Deficit entrainement
      Ligne 10 : Deficit repos
      Ligne 11 : Calories minimum absolues
      Ligne 12 : Repas par jour
      Lignes 15-20 : Glucides par seance (min/max)
    """
    def _get_float(row, col, defaut=0.0):
        val = ws_nutrition.cell(row=row, column=col).value
        try:
            return float(val)
        except (TypeError, ValueError):
            return defaut

    glucides = {}
    for row in range(15, 21):
        type_seance = ws_nutrition.cell(row=row, column=1).value
        if type_seance:
            type_seance = str(type_seance).strip()
            glucides[type_seance] = {
                "min_g_kg" : _get_float(row, 2, 3.0),
                "max_g_kg" : _get_float(row, 3, 5.0),
                "row"      : row,
            }

    return {
        "proteines_min"      : _get_float(4,  2, 1.8),
        "proteines_max"      : _get_float(5,  2, 2.5),
        "lipides_min"        : _get_float(7,  2, 0.8),
        "deficit_entrainement": _get_float(9,  2, 200.0),
        "deficit_repos"      : _get_float(10, 2, 0.0),
        "calories_min"       : _get_float(11, 2, 1800.0),
        "repas_par_jour"     : _get_float(12, 2, 3),
        "glucides_par_seance": glucides,
    }


# ------------------------------------------------------------
# CALCUL DES AJUSTEMENTS
# ------------------------------------------------------------
def _calculer_ajustements(
    statut   : str,
    params   : dict,
    actuels  : dict,
    objectif : str,
) -> dict:
    """
    Calcule les ajustements a appliquer selon le statut
    du suivi poids et l objectif de l athlete.

    Statuts possibles :
      perte_trop_rapide  -> reduire deficit, augmenter glucides
      gain_trop_rapide   -> reduire surplus, reduire glucides
      stagnation         -> augmenter deficit/surplus
      optimal            -> aucun ajustement
      sous_optimal       -> ajustement leger
      stable             -> aucun ajustement (maintien/perf)
      derive             -> correction vers le centre

    Retourne :
    {
        "deficit_entrainement_nouveau" : float,
        "deficit_repos_nouveau"        : float,
        "glucides_ajustes"             : {type_seance: {min, max}},
        "ajustements_appliques"        : [liste des changements],
        "aucun_ajustement"             : bool,
    }
    """
    direction  = params["direction"]
    ajustements_appliques = []

    # Copier les valeurs actuelles
    deficit_entrainement = actuels["deficit_entrainement"]
    deficit_repos        = actuels["deficit_repos"]
    glucides_ajustes     = {
        k: {"min_g_kg": v["min_g_kg"], "max_g_kg": v["max_g_kg"], "row": v["row"]}
        for k, v in actuels["glucides_par_seance"].items()
    }

    aucun_ajustement = False

    # ----------------------------------------------------------
    # PERTE TROP RAPIDE
    # Risque de perte musculaire -> reduire deficit
    # Source : Helms et al. 2014 / Garthe et al. 2011
    # ----------------------------------------------------------
    if statut == "perte_trop_rapide":
        ajust_def = params["ajustement_deficit_rapide"]
        ajust_glu = params["ajustement_glucides_rapide"]

        nouveau_deficit = max(
            deficit_entrainement - ajust_def,
            params["deficit_min_absolu"]
        )
        ajustements_appliques.append(
            f"Deficit entrainement reduit de {ajust_def} kcal "
            f"({deficit_entrainement:.0f} -> {nouveau_deficit:.0f} kcal) "
            f"— Perte trop rapide, risque de perte musculaire "
            f"(Helms et al. 2014)"
        )
        deficit_entrainement = nouveau_deficit

        # Augmenter les glucides pour proteger la masse musculaire
        for seance, val in glucides_ajustes.items():
            nouveau_min = round(val["min_g_kg"] * (1 + ajust_glu), 2)
            nouveau_max = round(val["max_g_kg"] * (1 + ajust_glu), 2)
            ajustements_appliques.append(
                f"Glucides {seance} augmentes de {ajust_glu*100:.0f}% "
                f"(min: {val['min_g_kg']} -> {nouveau_min} g/kg) "
                f"— Protection masse musculaire (Burke et al. 2011)"
            )
            glucides_ajustes[seance]["min_g_kg"] = nouveau_min
            glucides_ajustes[seance]["max_g_kg"] = nouveau_max

    # ----------------------------------------------------------
    # GAIN TROP RAPIDE
    # Accumulation excessive de graisse -> reduire surplus
    # Source : Antonio et al. 2020 ISSN / Haff & Triplett 2016
    # ----------------------------------------------------------
    elif statut == "gain_trop_rapide":
        ajust_def = params["ajustement_deficit_rapide"]
        ajust_glu = params["ajustement_glucides_rapide"]

        nouveau_deficit = min(
            deficit_entrainement + ajust_def,
            params["deficit_max_absolu"]
        )
        ajustements_appliques.append(
            f"Surplus reduit de {ajust_def} kcal "
            f"({deficit_entrainement:.0f} -> {nouveau_deficit:.0f} kcal) "
            f"— Gain trop rapide, risque d accumulation graisseuse "
            f"(Antonio et al. 2020)"
        )
        deficit_entrainement = nouveau_deficit

        # Reduire les glucides moderement
        for seance, val in glucides_ajustes.items():
            nouveau_min = round(val["min_g_kg"] * (1 + ajust_glu), 2)
            nouveau_max = round(val["max_g_kg"] * (1 + ajust_glu), 2)
            glucides_ajustes[seance]["min_g_kg"] = max(nouveau_min, 2.0)
            glucides_ajustes[seance]["max_g_kg"] = max(nouveau_max, 3.0)
        ajustements_appliques.append(
            f"Glucides reduits de {abs(ajust_glu)*100:.0f}% "
            f"— Limitation prise de graisse (Burke et al. 2011)"
        )

    # ----------------------------------------------------------
    # STAGNATION
    # Pas de progression -> augmenter deficit ou surplus
    # Source : Hall et al. 2012 — adaptation metabolique
    # ----------------------------------------------------------
    elif statut == "stagnation":
        ajust_def = params["ajustement_deficit_stagnation"]
        ajust_glu = params["ajustement_glucides_stagnation"]

        if direction == "perte":
            # Augmenter le deficit
            nouveau_deficit = min(
                deficit_entrainement + ajust_def,
                params["deficit_max_absolu"]
            )
            ajustements_appliques.append(
                f"Deficit entrainement augmente de {ajust_def} kcal "
                f"({deficit_entrainement:.0f} -> {nouveau_deficit:.0f} kcal) "
                f"— Stagnation detectee, relance de la perte "
                f"(Hall et al. 2012)"
            )
            deficit_entrainement = nouveau_deficit

            # Reduire les glucides repos
            for seance in ["Repos", "Recuperation"]:
                if seance in glucides_ajustes:
                    val = glucides_ajustes[seance]
                    nouveau_min = round(val["min_g_kg"] * (1 + ajust_glu), 2)
                    nouveau_max = round(val["max_g_kg"] * (1 + ajust_glu), 2)
                    glucides_ajustes[seance]["min_g_kg"] = max(nouveau_min, 1.5)
                    glucides_ajustes[seance]["max_g_kg"] = max(nouveau_max, 2.0)
            ajustements_appliques.append(
                f"Glucides repos reduits de {abs(ajust_glu)*100:.0f}% "
                f"— Amplification du deficit calorique les jours de repos "
                f"(Jeukendrup 2014)"
            )

        elif direction == "gain":
            # Augmenter le surplus
            nouveau_deficit = max(
                deficit_entrainement - ajust_def,
                -params["surplus_max_absolu"]
            )
            ajustements_appliques.append(
                f"Surplus augmente de {ajust_def} kcal "
                f"({deficit_entrainement:.0f} -> {nouveau_deficit:.0f} kcal) "
                f"— Stagnation, relance de la prise de masse "
                f"(Antonio et al. 2020)"
            )
            deficit_entrainement = nouveau_deficit

            # Augmenter les glucides entrainement
            for seance in ["Musculation", "Sprint", "Resistance"]:
                if seance in glucides_ajustes:
                    val = glucides_ajustes[seance]
                    nouveau_min = round(val["min_g_kg"] * (1 + ajust_glu), 2)
                    nouveau_max = round(val["max_g_kg"] * (1 + ajust_glu), 2)
                    glucides_ajustes[seance]["min_g_kg"] = nouveau_min
                    glucides_ajustes[seance]["max_g_kg"] = nouveau_max
            ajustements_appliques.append(
                f"Glucides entrainement augmentes de {ajust_glu*100:.0f}% "
                f"— Soutien anabolique (Slater & Phillips 2011)"
            )

    # ----------------------------------------------------------
    # OPTIMAL
    # Aucun ajustement necessaire
    # ----------------------------------------------------------
    elif statut == "optimal":
        aucun_ajustement = True
        ajustements_appliques.append(
            "Progression optimale — aucun ajustement necessaire"
        )

    # ----------------------------------------------------------
    # SOUS OPTIMAL
    # Ajustement tres leger
    # ----------------------------------------------------------
    elif statut == "sous_optimal":
        ajust_def = params["ajustement_deficit_stagnation"] // 2
        if direction == "perte":
            nouveau_deficit = min(
                deficit_entrainement + ajust_def,
                params["deficit_max_absolu"]
            )
            if nouveau_deficit != deficit_entrainement:
                ajustements_appliques.append(
                    f"Deficit ajuste legerement : "
                    f"{deficit_entrainement:.0f} -> {nouveau_deficit:.0f} kcal"
                )
                deficit_entrainement = nouveau_deficit

    # ----------------------------------------------------------
    # DERIVE (maintien/performance)
    # Correction vers la neutralite
    # Source : Loucks et al. 2011
    # ----------------------------------------------------------
    elif statut == "derive":
        ajust_def = params["ajustement_deficit_stagnation"] // 2
        # Ramener vers le centre
        if deficit_entrainement > params["deficit_max_absolu"] / 2:
            deficit_entrainement = max(
                deficit_entrainement - ajust_def,
                params["deficit_min_absolu"]
            )
            ajustements_appliques.append(
                f"Deficit reduit pour stabilisation "
                f"(Loucks et al. 2011)"
            )
        elif deficit_entrainement < params["deficit_min_absolu"]:
            deficit_entrainement = min(
                deficit_entrainement + ajust_def,
                params["deficit_max_absolu"]
            )
            ajustements_appliques.append(
                f"Deficit augmente pour stabilisation"
            )

    # ----------------------------------------------------------
    # STABLE (maintien/performance)
    # ----------------------------------------------------------
    elif statut == "stable":
        aucun_ajustement = True
        ajustements_appliques.append(
            "Poids stable — aucun ajustement necessaire"
        )

    return {
        "deficit_entrainement_nouveau" : round(deficit_entrainement, 1),
        "deficit_repos_nouveau"        : round(deficit_repos, 1),
        "glucides_ajustes"             : glucides_ajustes,
        "ajustements_appliques"        : ajustements_appliques,
        "aucun_ajustement"             : aucun_ajustement,
    }


# ------------------------------------------------------------
# ECRITURE DES AJUSTEMENTS DANS EXCEL
# ------------------------------------------------------------
def _ecrire_ajustements_excel(
    ajustements : dict,
    actuels     : dict,
) -> None:
    """
    Ecrit les nouveaux parametres dans la feuille NUTRITION
    de athlete_config.xlsm.

    Ne modifie que les cellules de valeurs numeriques.
    Ne touche jamais aux labels ni aux autres feuilles.
    """
    # Verifier que le fichier n est pas ouvert
    try:
        with open(EXCEL_CONFIG_PATH, "a"):
            pass
    except PermissionError:
        raise PermissionError(
            f"[ERREUR] Fermez athlete_config.xlsm dans Excel avant "
            f"de lancer l adaptation automatique."
        )

    wb            = openpyxl.load_workbook(EXCEL_CONFIG_PATH)
    ws_nutrition  = wb["NUTRITION"]

    # Mettre a jour deficit entrainement (ligne 9, col 2)
    ws_nutrition.cell(row=9, column=2).value = (
        ajustements["deficit_entrainement_nouveau"]
    )

    # Mettre a jour deficit repos (ligne 10, col 2)
    ws_nutrition.cell(row=10, column=2).value = (
        ajustements["deficit_repos_nouveau"]
    )

    # Mettre a jour les glucides par seance
    for seance, val in ajustements["glucides_ajustes"].items():
        row = val.get("row")
        if row:
            ws_nutrition.cell(row=row, column=2).value = val["min_g_kg"]
            ws_nutrition.cell(row=row, column=3).value = val["max_g_kg"]

    # Ajouter un commentaire d adaptation dans une cellule dediee
    date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws_nutrition.cell(row=25, column=1).value = (
        f"Derniere adaptation : {date_str}"
    )

    wb.save(EXCEL_CONFIG_PATH)
    wb.close()
    print(f"[OK] Parametres mis a jour dans m")


# ------------------------------------------------------------
# FONCTION PRINCIPALE
# ------------------------------------------------------------
def adapter_parametres(
    analyse_poids : dict,
    config        : dict,
    appliquer     : bool = True,
) -> dict:
    """
    Fonction principale d adaptation automatique.

    1. Analyse le statut du suivi poids
    2. Calcule les ajustements necessaires
    3. Applique les ajustements dans Excel (si appliquer=True)
    4. Retourne un rapport complet

    Parametre appliquer :
      True  -> ecrit dans Excel (mode production)
      False -> calcule sans ecrire (mode simulation)
    """
    statut  = analyse_poids.get("statut", "optimal")
    params  = analyse_poids.get("params", {})
    objectif = config["profil"]["objectif"]

    print(f"\n[...] Analyse adaptation nutritionnelle...")
    print(f"      Statut poids   : {statut.upper()}")
    print(f"      Objectif       : {objectif}")

    if statut == "pas_de_donnees":
        print("[INFO] Pas de donnees de poids disponibles.")
        print("[INFO] Renseignez la feuille SUIVI_POIDS.")
        return {
            "statut"               : "pas_de_donnees",
            "ajustements_appliques": [],
            "aucun_ajustement"     : True,
             "applique"            : False
        }

    # Lire les parametres actuels depuis Excel
    wb           = openpyxl.load_workbook(EXCEL_CONFIG_PATH, data_only=True)
    ws_nutrition = wb["NUTRITION"]
    actuels      = _lire_parametres_actuels(ws_nutrition)
    wb.close()

    # Calculer les ajustements
    ajustements = _calculer_ajustements(
        statut   = statut,
        params   = params,
        actuels  = actuels,
        objectif = objectif,
    )

    # Afficher les ajustements
    print(f"\n  Ajustements calcules :")
    for msg in ajustements["ajustements_appliques"]:
        print(f"    -> {msg}")

    if not ajustements["aucun_ajustement"]:
        print(f"\n  Nouveaux parametres :")
        print(
            f"    Deficit entrainement : "
            f"{actuels['deficit_entrainement']:.0f} kcal "
            f"-> {ajustements['deficit_entrainement_nouveau']:.0f} kcal"
        )
        print(
            f"    Deficit repos        : "
            f"{actuels['deficit_repos']:.0f} kcal "
            f"-> {ajustements['deficit_repos_nouveau']:.0f} kcal"
        )

        # Appliquer si demande
        if appliquer:
            _ecrire_ajustements_excel(ajustements, actuels)
        else:
            print(f"\n  [SIMULATION] Ajustements calcules mais non appliques.")
            print(f"  [INFO] Passez appliquer=True pour ecrire dans Excel.")

    return {
        "statut"                       : statut,
        "objectif"                     : objectif,
        "parametres_avant"             : actuels,
        "deficit_entrainement_nouveau" : ajustements["deficit_entrainement_nouveau"],
        "deficit_repos_nouveau"        : ajustements["deficit_repos_nouveau"],
        "glucides_ajustes"             : ajustements["glucides_ajustes"],
        "ajustements_appliques"        : ajustements["ajustements_appliques"],
        "aucun_ajustement"             : ajustements["aucun_ajustement"],
        "applique"                     : appliquer and not ajustements["aucun_ajustement"],
    }


# ------------------------------------------------------------
# TEST AUTONOME
# ------------------------------------------------------------
if __name__ == "__main__":
    sys.path.insert(0, os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "..")
    ))

    from optimisation.excel.reader import lire_config_athlete
    from suivi.tracking.poids      import analyser_poids

    config  = lire_config_athlete()
    analyse = analyser_poids(config)

    rapport = adapter_parametres(
        analyse_poids = analyse,
        config        = config,
        appliquer     = False,
    )

    print(f"\n[OK] Simulation terminee.")
    print(f"     Statut      : {rapport['statut']}")
    print(f"     Ajustements : {len(rapport['ajustements_appliques'])}")
    print(f"     Applique    : {rapport['applique']}")
