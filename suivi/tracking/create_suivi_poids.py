# ============================================================
# CREATION DE LA FEUILLE SUIVI_POIDS
# suivi/tracking/create_suivi_poids.py
# ============================================================

import os
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
sys.dont_write_bytecode = True

sys.path.insert(0, os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..")
))
from suivi.config import EXCEL_CONFIG_PATH


def create_feuille_suivi_poids():

    wb = openpyxl.load_workbook(EXCEL_CONFIG_PATH)

    # Supprimer si existe deja
    if "SUIVI_POIDS" in wb.sheetnames:
        del wb["SUIVI_POIDS"]

    ws = wb.create_sheet("SUIVI_POIDS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "E74C3C"

    # Styles
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def font(hex_color, bold=False, size=10):
        return Font(color=hex_color, bold=bold, size=size, name="Calibri")
    def align(h="left", v="center"):
        return Alignment(horizontal=h, vertical=v)
    def border():
        side = Side(style="thin", color="B0BEC5")
        return Border(left=side, right=side, top=side, bottom=side)

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 14   # Date
    ws.column_dimensions["B"].width = 12   # Poids
    ws.column_dimensions["C"].width = 12   # Variation
    ws.column_dimensions["D"].width = 12   # Objectif
    ws.column_dimensions["E"].width = 25   # Statut
    ws.column_dimensions["F"].width = 30   # Notes

    # Titre
    ws.merge_cells("A1:F1")
    c           = ws["A1"]
    c.value     = "SUIVI DU POIDS HEBDOMADAIRE"
    c.fill      = fill("2E4057")
    c.font      = font("FFFFFF", bold=True, size=14)
    c.alignment = align("left", "center")
    ws.row_dimensions[1].height = 32

    # Sous-titre
    ws.merge_cells("A2:F2")
    c           = ws["A2"]
    c.value     = (
        "Renseignez votre poids chaque semaine. "
        "Le systeme ajuste automatiquement vos parametres nutritionnels."
    )
    c.fill      = fill("C8972B")
    c.font      = font("FFFFFF", size=9)
    c.alignment = align("left", "center")
    ws.row_dimensions[2].height = 16

    # Header
    headers = ["Date", "Poids (kg)", "Variation (kg)", "Objectif (kg)", "Statut", "Notes"]
    for col, titre in enumerate(headers, 1):
        c           = ws.cell(row=3, column=col)
        c.value     = titre
        c.fill      = fill("1C2B3A")
        c.font      = font("FFFFFF", bold=True, size=9)
        c.alignment = align("center", "center")
        c.border    = border()
    ws.row_dimensions[3].height = 20

    # Ligne exemple avec poids initial
    date_debut = datetime.now().strftime("%d/%m/%Y")
    exemples = [
        (date_debut, 80.0, "", 73.0, "Debut du programme", "Poids de depart"),
        ("",         "",   "", "",   "",                   ""),
        ("",         "",   "", "",   "",                   ""),
        ("",         "",   "", "",   "",                   ""),
        ("",         "",   "", "",   "",                   ""),
    ]

    for i, (date, poids, var, obj, statut, notes) in enumerate(exemples, start=4):
        alt = (i % 2 == 0)
        bg  = "EEF2F5" if alt else "FFFFFF"

        valeurs = [date, poids, var, obj, statut, notes]
        for col, val in enumerate(valeurs, 1):
            c           = ws.cell(row=i, column=col)
            c.value     = val
            c.fill      = fill(bg)
            c.font      = font("2C2C2C", size=10)
            c.alignment = align("center" if col != 6 else "left", "center")
            c.border    = border()
        ws.row_dimensions[i].height = 20

    # Note bas de page
    ws.merge_cells("A10:F10")
    c           = ws["A10"]
    c.value     = (
        "Format date : JJ/MM/AAAA — "
        "Pesee recommandee : le matin a jeun, meme conditions chaque semaine."
    )
    c.font      = font("1C2B3A", size=8)
    c.alignment = align("left", "center")
    ws.row_dimensions[10].height = 14

    wb.save(EXCEL_CONFIG_PATH)
    wb.close()
    print(f"[OK] Feuille SUIVI_POIDS creee dans : {EXCEL_CONFIG_PATH}")
    print(f"[INFO] Renseignez votre poids chaque semaine dans la colonne B")


if __name__ == "__main__":
    create_feuille_suivi_poids()
