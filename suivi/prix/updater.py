# ============================================================
# MISE A JOUR DES PRIX EN BASE
# suivi/prix/updater.py
#
# Lit le fichier prix.xlsx et met a jour
# la colonne cout_kg dans nutrition.db
# pour chaque aliment renseigne.
#
# Calcule egalement le prix moyen historique
# et le promo_score pour chaque aliment.
# ============================================================

import os
import sys
import sqlite3
import openpyxl
from datetime import datetime
sys.dont_write_bytecode = True

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import (
    NUTRITION_DB,
    PRIX_XLSX,
    NB_SEMAINES_HISTORIQUE_PRIX,
    PROMO_SCORE_MIN,
)


# ------------------------------------------------------------
# CONNEXION BASE DE DONNEES
# ------------------------------------------------------------
def _get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(NUTRITION_DB)
    conn.row_factory = sqlite3.Row
    return conn


# ------------------------------------------------------------
# CREATION TABLE HISTORIQUE PRIX
# Creee si elle n existe pas encore
# ------------------------------------------------------------
def _create_table_prix(conn: sqlite3.Connection) -> None:
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS historique_prix (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            aliment_id  INTEGER NOT NULL,
            nom_aliment TEXT,
            magasin     TEXT,
            prix_actuel REAL NOT NULL,
            prix_moyen  REAL,
            promo_score REAL DEFAULT 0.0,
            date_releve TEXT NOT NULL,
            semaine     INTEGER,
            annee       INTEGER,
            FOREIGN KEY (aliment_id) REFERENCES aliments(id)
        )
    """)
    conn.commit()


# ------------------------------------------------------------
# LECTURE DU FICHIER PRIX.XLSX
# ------------------------------------------------------------
def _lire_prix_excel() -> list:
    """
    Lit le fichier prix.xlsx et retourne une liste
    de dictionnaires avec les prix saisis.

    Format attendu dans prix.xlsx :
      Feuille PRIX_SEMAINE :
        Col A : Nom aliment (doit correspondre a nutrition.db)
        Col B : Magasin
        Col C : Prix actuel (euros/kg)
        Col D : Notes (optionnel)

    Retourne :
      [
        {
          "nom"        : "Poulet blanc cuit",
          "magasin"    : "Leclerc",
          "prix"       : 8.50,
          "notes"      : "",
        },
        ...
      ]
    """
    if not os.path.exists(PRIX_XLSX):
        print(f"[WARN] Fichier prix introuvable : {PRIX_XLSX}")
        print(f"[INFO] Creez le fichier via : python suivi/prix/create_prix_excel.py")
        return []

    wb     = openpyxl.load_workbook(PRIX_XLSX, data_only=True)

    if "PRIX_SEMAINE" not in wb.sheetnames:
        print(f"[WARN] Feuille 'PRIX_SEMAINE' introuvable dans prix.xlsx")
        return []

    ws     = wb["PRIX_SEMAINE"]
    prix   = []
    row    = 4   # Les donnees commencent ligne 4 (apres titre + header)

    while ws.cell(row=row, column=1).value:
        nom     = str(ws.cell(row=row, column=1).value or "").strip()
        magasin = str(ws.cell(row=row, column=2).value or "").strip()
        prix_v  = ws.cell(row=row, column=3).value
        notes   = str(ws.cell(row=row, column=4).value or "").strip()

        if nom and prix_v is not None:
            try:
                prix.append({
                    "nom"     : nom,
                    "magasin" : magasin,
                    "prix"    : float(prix_v),
                    "notes"   : notes,
                })
            except (ValueError, TypeError):
                print(f"[WARN] Prix invalide ligne {row} : {prix_v}")

        row += 1

    wb.close()
    print(f"[OK] {len(prix)} prix lus depuis prix.xlsx")
    return prix


# ------------------------------------------------------------
# RECHERCHE D UN ALIMENT EN BASE
# ------------------------------------------------------------
def _trouver_aliment(conn: sqlite3.Connection, nom: str) -> dict | None:
    """
    Cherche un aliment dans nutrition.db par son nom.
    Recherche exacte d abord, puis partielle.
    """
    cursor = conn.cursor()

    # Recherche exacte
    cursor.execute(
        "SELECT id, nom, cout_kg FROM aliments WHERE LOWER(nom) = LOWER(?)",
        (nom,)
    )
    row = cursor.fetchone()
    if row:
        return dict(row)

    # Recherche partielle
    cursor.execute(
        "SELECT id, nom, cout_kg FROM aliments WHERE LOWER(nom) LIKE LOWER(?)",
        (f"%{nom}%",)
    )
    row = cursor.fetchone()
    if row:
        return dict(row)

    return None


# ------------------------------------------------------------
# CALCUL DU PRIX MOYEN HISTORIQUE
# ------------------------------------------------------------
def _calculer_prix_moyen(
    conn       : sqlite3.Connection,
    aliment_id : int,
    prix_actuel: float,
) -> float:
    """
    Calcule le prix moyen sur les NB_SEMAINES_HISTORIQUE_PRIX
    dernieres semaines.

    Si pas d historique -> retourne le prix actuel.
    """
    cursor = conn.cursor()
    cursor.execute("""
        SELECT AVG(prix_actuel) as moy
        FROM historique_prix
        WHERE aliment_id = ?
        ORDER BY date_releve DESC
        LIMIT ?
    """, (aliment_id, NB_SEMAINES_HISTORIQUE_PRIX))

    row = cursor.fetchone()
    if row and row["moy"] is not None:
        # Moyenne ponderee : 70% historique + 30% prix actuel
        return round(row["moy"] * 0.7 + prix_actuel * 0.3, 3)

    return prix_actuel


# ------------------------------------------------------------
# CALCUL DU PROMO SCORE
# ------------------------------------------------------------
def _calculer_promo_score(prix_actuel: float, prix_moyen: float) -> float:
    """
    promo_score = (prix_moyen - prix_actuel) / prix_moyen

    Interpretation :
      > 0.15  : promotion interessante (>15% de reduction)
      0 - 0.15: prix normal
      < 0     : prix plus eleve que d habitude

    Source : logique metier interne
    """
    if prix_moyen <= 0:
        return 0.0
    return round((prix_moyen - prix_actuel) / prix_moyen, 4)


# ------------------------------------------------------------
# MISE A JOUR D UN ALIMENT
# ------------------------------------------------------------
def _mettre_a_jour_aliment(
    conn        : sqlite3.Connection,
    aliment_id  : int,
    nom_aliment : str,
    magasin     : str,
    prix_actuel : float,
) -> dict:
    """
    Met a jour le cout_kg dans aliments
    et insere un enregistrement dans historique_prix.

    Retourne un dictionnaire avec les details de la mise a jour.
    """
    cursor     = conn.cursor()
    maintenant = datetime.now()
    semaine    = maintenant.isocalendar()[1]
    annee      = maintenant.year

    # Verifier si un releve existe deja cette semaine
    cursor.execute("""
        SELECT id FROM historique_prix
        WHERE aliment_id = ?
          AND semaine    = ?
          AND annee      = ?
    """, (aliment_id, semaine, annee))

    existe_deja = cursor.fetchone()

    # Calcul prix moyen et promo score
    prix_moyen  = _calculer_prix_moyen(conn, aliment_id, prix_actuel)
    promo_score = _calculer_promo_score(prix_actuel, prix_moyen)

    if existe_deja:
        # Mettre a jour l enregistrement existant
        cursor.execute("""
            UPDATE historique_prix
            SET prix_actuel = ?,
                prix_moyen  = ?,
                promo_score = ?,
                magasin     = ?,
                date_releve = ?
            WHERE id = ?
        """, (
            prix_actuel, prix_moyen, promo_score,
            magasin, maintenant.strftime("%Y-%m-%d %H:%M:%S"),
            existe_deja["id"]
        ))
    else:
        # Inserer un nouvel enregistrement
        cursor.execute("""
            INSERT INTO historique_prix (
                aliment_id, nom_aliment, magasin,
                prix_actuel, prix_moyen, promo_score,
                date_releve, semaine, annee
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            aliment_id, nom_aliment, magasin,
            prix_actuel, prix_moyen, promo_score,
            maintenant.strftime("%Y-%m-%d %H:%M:%S"),
            semaine, annee
        ))

    # Mettre a jour le cout_kg dans la table aliments
    cursor.execute("""
        UPDATE aliments
        SET cout_kg = ?
        WHERE id = ?
    """, (prix_actuel, aliment_id))

    conn.commit()

    return {
        "aliment_id"  : aliment_id,
        "nom"         : nom_aliment,
        "magasin"     : magasin,
        "prix_actuel" : prix_actuel,
        "prix_moyen"  : prix_moyen,
        "promo_score" : promo_score,
        "en_promo"    : promo_score >= PROMO_SCORE_MIN,
    }


# ------------------------------------------------------------
# NETTOYAGE HISTORIQUE ANCIEN
# ------------------------------------------------------------
def _nettoyer_historique(conn: sqlite3.Connection) -> None:
    """
    Supprime les releves de prix plus anciens que
    NB_SEMAINES_HISTORIQUE_PRIX semaines.
    """
    cursor = conn.cursor()
    cursor.execute("""
        DELETE FROM historique_prix
        WHERE date_releve < datetime('now', ? || ' weeks')
    """, (f"-{NB_SEMAINES_HISTORIQUE_PRIX}",))
    nb_supprimes = cursor.rowcount
    conn.commit()

    if nb_supprimes > 0:
        print(f"[OK] {nb_supprimes} releves anciens supprimes de l historique")


# ------------------------------------------------------------
# FONCTION PRINCIPALE
# ------------------------------------------------------------
def mettre_a_jour_prix() -> list:
    """
    Fonction principale de mise a jour des prix.

    Lit prix.xlsx -> met a jour nutrition.db
    Retourne la liste des mises a jour effectuees.
    """
    print("\n[...] Mise a jour des prix alimentaires...")

    conn = _get_connection()

    # Creer la table si elle n existe pas
    _create_table_prix(conn)

    # Nettoyer l historique ancien
    _nettoyer_historique(conn)

    # Lire les prix depuis Excel
    prix_excel = _lire_prix_excel()

    if not prix_excel:
        print("[WARN] Aucun prix a mettre a jour.")
        conn.close()
        return []

    resultats    = []
    mis_a_jour   = 0
    non_trouves  = []
    en_promo     = []

    for entree in prix_excel:
        nom         = entree["nom"]
        magasin     = entree["magasin"]
        prix_actuel = entree["prix"]

        # Chercher l aliment en base
        aliment = _trouver_aliment(conn, nom)

        if aliment is None:
            non_trouves.append(nom)
            continue

        # Mettre a jour
        resultat = _mettre_a_jour_aliment(
            conn,
            aliment["id"],
            aliment["nom"],
            magasin,
            prix_actuel,
        )

        resultats.append(resultat)
        mis_a_jour += 1

        if resultat["en_promo"]:
            en_promo.append(resultat)

    conn.close()

    # Rapport
    print(f"\n[OK] Prix mis a jour : {mis_a_jour}")

    if en_promo:
        print(f"\n  Promotions detectees ({len(en_promo)}) :")
        for r in sorted(en_promo, key=lambda x: x["promo_score"], reverse=True):
            print(
                f"    {r['nom']:<40}"
                f" {r['prix_actuel']:.2f} e/kg"
                f" (moy: {r['prix_moyen']:.2f} e/kg)"
                f" -> -{r['promo_score']*100:.0f}%"
            )

    if non_trouves:
        print(f"\n  Aliments non trouves en base ({len(non_trouves)}) :")
        for nom in non_trouves:
            print(f"    -> {nom}")
        print(f"  [INFO] Verifiez l orthographe dans prix.xlsx")

    return resultats


# ------------------------------------------------------------
# TEST AUTONOME
# ------------------------------------------------------------
if __name__ == "__main__":
    resultats = mettre_a_jour_prix()
    print(f"\n  Total traite : {len(resultats)} aliments")